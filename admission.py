from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import streamlit as st
import pandas as pd
import joblib
from io import BytesIO

def main():
    # data = {
    #     'choice1': ['BSN', 'BSCS', 'BSA', 'BSIT'],
    #     'choice2': ['BSCS', 'BSN', 'BSCE', 'BSCS'],
    #     'choice3': ['BSPT', 'BSA', 'BSPT', 'BSPT'],
    #     'strand': ['STEM', 'ABM', 'TVL', 'STEM'],
    #     'gwa': [90.5, 88.0, 92.3, 85.6],
    #     'rating': [88.0, 89.2, 88.2, 89.3]
    # }

    # df = pd.DataFrame(data)
    # print(df)

    # Set page title and description
    st.markdown("""
    <style>
    .title {
        text-align: center;
    }
    </style>
    """, unsafe_allow_html=True)

    if 'file' not in st.session_state:
        st.session_state.file = None
    if 'file_name' not in st.session_state:
        st.session_state.file_name = None

    st.markdown("<h1 class='title'>PLMAT Course Recommendation</h1>", unsafe_allow_html=True)

    uploaded_file = st.file_uploader("Choose an Excel file", type=['xlsx'])
    
    if uploaded_file is not None:
        if uploaded_file.file_id != st.session_state.file_name:
            st.session_state.file = None
        
        if st.session_state.file is None:
            st.session_state.file_name = uploaded_file.file_id
            df = pd.read_excel(uploaded_file)
            st.dataframe(df, use_container_width=True)

            df_new = df.copy().drop(columns=['program'])
            predict(df_new)
            # st.dataframe(df['PREDICTION'], use_container_width=True)
            st.dataframe(df_new['prediction'], use_container_width=True)

            df['prediction'] = df_new['prediction']

            buffer = BytesIO()
            df.to_excel(buffer, index=False, engine='openpyxl') 
            buffer.seek(0)

            # Save the buffer to a temporary file to apply auto_fit_columns
            temp_file_path = 'temp_predicted_courses.xlsx'
            with open(temp_file_path, 'wb') as temp_file:
                temp_file.write(buffer.getbuffer())

            # Apply auto_fit_columns to the temporary file
            auto_fit_columns(temp_file_path)

            # Read the adjusted file back into the buffer
            with open(temp_file_path, 'rb') as temp_file:
                buffer = BytesIO(temp_file.read())

            st.download_button(
                label="Download",
                data=buffer,
                file_name='Predicted_Courses.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                use_container_width=True
            )
    else:
        st.write("Please upload a file to proceed")

def predict(df):
    # Load the Random Forest model
    random_forest_model = joblib.load('random_forest_model.pkl')

    # Load the LabelEncoder
    choice1_encoder = joblib.load('choice1_encoder.pkl')
    choice2_encoder = joblib.load('choice2_encoder.pkl')
    choice3_encoder = joblib.load('choice3_encoder.pkl')
    strand_encoder = joblib.load('strand_encoder.pkl')

    course_mapping = {
        'BAC': 'BAC',
        'BPA': 'Bachelor of Public Administration',
        'BSA': 'Bachelor of Science in Accountancy',
        'BSArch': 'Bachelor of Science in Architecture',
        'BSBABE': 'Bachelor of Science in Business Administration Major in Business Economics',
        'BSBAFM': 'Bachelor of Science in Business Administration Major in Finance Management',
        'BSBAHRM': 'Bachelor of Science in Business Administration Major in Hotel and Restaurant Management',
        'BSBAMM': 'Bachelor of Science in Business Administration Major in Marketing Management',
        'BSBio': 'Bachelor of Science in Biology',
        'BSCE': 'Bachelor of Science in Civil Engineering',
        'BSCHE': 'Bachelor of Science in Chemical Engineering',
        'BSCS': 'Bachelor of Science in Computer Science',
        'BSChem': 'Bachelor of Science in Chemistry',
        'BSCpE': 'Bachelor of Science in Computer Engineering',
        'BSECE': 'Bachelor of Science in Electronics Engineering',
        'BSEE': 'Bachelor of Science in Electrical Engineering',
        'BSEd-Eng': 'Bachelor of Science in Education - English',
        'BSEd-Fil': 'Bachelor of Science in Education - Filipino',
        'BSEd-Math': 'Bachelor of Science in Education - Math',
        'BSEd-SS': 'Bachelor of Science in Education - Social Studies',
        'BSEd-Sci': 'Bachelor of Science in Education - Science',
        'BSEntrep': 'Bachelor of Science in Entrepreneurship',
        'BSHM': 'Bachelor of Science in Hospitality Management',
        'BSIT': 'Bachelor of Science in Information Technology',
        'BSME': 'Bachelor of Science in Mechanical Engineering',
        'BSMath': 'Bachelor of Science in Math',
        'BSN': 'Bachelor of Science in Nursing',
        'BSPSY': 'Bachelor of Science in Psychology',
        'BSPT': 'Bachelor of Science in Physical Therapy',
        'BSREM': 'Bachelor of Science in Real Estate Management',
        'BSSW': 'Bachelor of Science in Social Work',
        'BSTM': 'Bachelor of Science in Tourism Management'
    }

    df['choice1'] = choice1_encoder.transform(df['choice1'])
    df['choice2'] = choice2_encoder.transform(df['choice2'])
    df['choice3'] = choice3_encoder.transform(df['choice3'])
    df['strand'] = strand_encoder.transform(df['strand'])
    df['prediction'] = random_forest_model.predict(df)
    # df.columns = [col.upper() for col in df.columns]
    df['prediction'] = df['prediction'].map(course_mapping)

def auto_fit_columns(file_path):
    workbook = load_workbook(file_path)
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        for col in worksheet.columns:
            max_length = 0
            column = get_column_letter(col[0].column)  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.1
            worksheet.column_dimensions[column].width = adjusted_width
    workbook.save(file_path)

if __name__ == '__main__':
    main()